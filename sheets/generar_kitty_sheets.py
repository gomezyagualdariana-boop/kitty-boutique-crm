# -*- coding: utf-8 -*-
"""
Generador de Kitty_Boutique_Sheets.xlsx para subir a Google Sheets.

Crea dos hojas:
  - "Leads": ID + columnas del CRM, con listas desplegables (data validation)
    en las columnas de categoría y formato condicional por Estado.
  - "Dashboard": KPIs y tablas auxiliares con FORMULAS (SUMIFS/COUNTIFS/
    SUMPRODUCT+TEXT), nunca valores fijos, alimentando gráficos nativos de Excel.

Requiere: pip install openpyxl
Ejecutar:  python generar_kitty_sheets.py
"""

import random
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import DoughnutChart, BarChart, Reference
from openpyxl.utils import get_column_letter

random.seed(42)

# ---------------- Estructura real basada en Kitty_Boutique_100_Clientes.xlsx ----------------
ESTADOS = ['Pendiente', 'Contactada', 'En negociación', 'Apartado', 'Vendido']
FUENTES = ['Página Web', 'Instagram', 'Facebook', 'TikTok', 'Referido', 'WhatsApp']
PRODUCTOS = ['Falda Jeans', 'Top Elegante', 'Accesorios', 'Conjunto Casual', 'Bolso Fashion',
             'Blazer Mujer', 'Vestido Floral', 'Blusa Casual', 'Pantalón Jeans', 'Blusa Suplex']
PRIORIDADES = ['Alta', 'Media', 'Baja']
MESES = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio']
EMPRESA = 'Kitty Boutique'

NOMBRES = ['Ana', 'Cinthya', 'Nicole', 'Verónica', 'María José', 'Gabriela', 'Daniela', 'Karla',
           'Fernanda', 'Valeria', 'Andrea', 'Camila', 'Lucía', 'Mishell', 'Paola', 'Ximena',
           'Briggitte', 'Yadira', 'Estefanía', 'Mayra', 'Johanna', 'Doménica', 'Melanie', 'Jazmín', 'Liseth']
APELLIDOS = ['Panchana', 'Mendoza', 'Rodríguez', 'Gómez', 'Alvarado', 'Suárez', 'Zambrano',
             'Cedeño', 'Vélez', 'Macías', 'Torres', 'Pincay', 'Loor', 'Quiñonez', 'Yagual',
             'Tomalá', 'Bravo', 'Reyes', 'Chávez', 'Borbor']

PRODUCT_NOTE = {
    'Falda Jeans': 'falda jeans', 'Top Elegante': 'top elegante', 'Accesorios': 'accesorios',
    'Conjunto Casual': 'conjunto casual', 'Bolso Fashion': 'bolso fashion', 'Blazer Mujer': 'blazer mujer',
    'Vestido Floral': 'vestido floral', 'Blusa Casual': 'blusa casual', 'Pantalón Jeans': 'pantalón jeans',
    'Blusa Suplex': 'blusa suplex'
}

CANTIDAD = 128

HEADERS = ['ID', 'Fecha', 'Cliente', 'Empresa', 'Producto', 'Valor USD', 'Estado', 'Fuente',
           'Prioridad', 'Mes', 'Notas', 'Fecha Seguimiento']

ESTADO_COLORS = {
    'Pendiente': 'FFF3DA', 'Contactada': 'DCEEFB', 'En negociación': 'FDE6C8',
    'Apartado': 'E8DEFB', 'Vendido': 'DDF5E6'
}

PINK = '00E91E8C'
PASTEL = '00FFD6E8'


def generar_filas(n):
    filas = []
    base = datetime(2026, 1, 1)
    for i in range(n):
        fecha = base + timedelta(days=int(i * (180 / n)) + random.randint(0, 2))
        seguimiento = fecha + timedelta(days=random.randint(3, 25))
        producto = random.choice(PRODUCTOS)
        estado = random.choice(ESTADOS)
        filas.append({
            'id': f'L{1000 + i}',
            'fecha': fecha.strftime('%d/%m/%Y'),
            'cliente': f'{random.choice(NOMBRES)} {random.choice(APELLIDOS)}',
            'empresa': EMPRESA,
            'producto': producto,
            'valor': round(random.uniform(6.5, 25), 2),
            'estado': estado,
            'fuente': random.choice(FUENTES),
            'prioridad': random.choice(PRIORIDADES),
            'mes': MESES[fecha.month - 1] if fecha.month <= 6 else MESES[(fecha.month - 1) % 6],
            'notas': f'Interesada en {PRODUCT_NOTE[producto]}',
            'fecha_seg': seguimiento.strftime('%d/%m/%Y')
        })
    return filas


def construir_hoja_leads(wb, filas):
    ws = wb.active
    ws.title = 'Leads'

    header_fill = PatternFill(start_color=PINK, end_color=PINK, fill_type='solid')
    header_font = Font(color='FFFFFFFF', bold=True)

    for c, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for r, f in enumerate(filas, start=2):
        ws.cell(row=r, column=1, value=f['id'])
        ws.cell(row=r, column=2, value=f['fecha'])
        ws.cell(row=r, column=3, value=f['cliente'])
        ws.cell(row=r, column=4, value=f['empresa'])
        ws.cell(row=r, column=5, value=f['producto'])
        ws.cell(row=r, column=6, value=f['valor'])
        ws.cell(row=r, column=7, value=f['estado'])
        ws.cell(row=r, column=8, value=f['fuente'])
        ws.cell(row=r, column=9, value=f['prioridad'])
        ws.cell(row=r, column=10, value=f['mes'])
        ws.cell(row=r, column=11, value=f['notas'])
        ws.cell(row=r, column=12, value=f['fecha_seg'])

    last_row = len(filas) + 1

    # Anchos de columna
    widths = [10, 12, 22, 16, 16, 11, 16, 14, 11, 10, 32, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---- Listas desplegables (data validation) ----
    def add_dv(col_letter, options):
        formula = '"' + ','.join(options) + '"'
        dv = DataValidation(type='list', formula1=formula, allow_blank=True, showDropDown=False)
        ws.add_data_validation(dv)
        dv.add(f'{col_letter}2:{col_letter}{max(last_row, 200)}')

    add_dv('E', PRODUCTOS)     # Producto
    add_dv('G', ESTADOS)       # Estado
    add_dv('H', FUENTES)       # Fuente
    add_dv('I', PRIORIDADES)   # Prioridad
    add_dv('J', MESES)         # Mes

    # ---- Formato condicional por Estado (columna G) ----
    for estado, color in ESTADO_COLORS.items():
        rule = CellIsRule(operator='equal', formula=[f'"{estado}"'],
                           fill=PatternFill(start_color='00' + color[2:], end_color='00' + color[2:], fill_type='solid'))
        ws.conditional_formatting.add(f'G2:G{max(last_row, 200)}', rule)

    ws.freeze_panes = 'A2'
    return ws, last_row


def construir_hoja_dashboard(wb, last_row):
    ws = wb.create_sheet('Dashboard')

    title_font = Font(size=16, bold=True, color='00C4126F')
    ws['A1'] = 'KITTY BOUTIQUE — DASHBOARD'
    ws['A1'].font = title_font

    label_font = Font(bold=True, color='008A6B78', size=10)
    kpi_font = Font(bold=True, size=18, color='00E91E8C')

    # ---- KPIs (fila 3) con fórmulas ----
    ws['A3'] = 'TOTAL LEADS'
    ws['A3'].font = label_font
    ws['A4'] = f'=COUNTA(Leads!A2:A{last_row})'
    ws['A4'].font = kpi_font

    ws['C3'] = 'VALOR DEL PIPELINE'
    ws['C3'].font = label_font
    ws['C4'] = f'=SUMIFS(Leads!F2:F{last_row},Leads!G2:G{last_row},"<>Vendido")'
    ws['C4'].font = kpi_font
    ws['C4'].number_format = '$#,##0.00'

    ws['E3'] = 'INGRESOS GANADOS'
    ws['E3'].font = label_font
    ws['E4'] = f'=SUMIFS(Leads!F2:F{last_row},Leads!G2:G{last_row},"Vendido")'
    ws['E4'].font = kpi_font
    ws['E4'].number_format = '$#,##0.00'

    ws['G3'] = 'TASA DE CONVERSIÓN'
    ws['G3'].font = label_font
    ws['G4'] = f'=COUNTIF(Leads!G2:G{last_row},"Vendido")/COUNTA(Leads!A2:A{last_row})'
    ws['G4'].font = kpi_font
    ws['G4'].number_format = '0.0%'

    # ---- Tablas auxiliares — alejadas hacia la derecha (a partir de columna N) ----
    # Tabla 1: por Estado (embudo) -> columnas N:O, filas 3..
    ws['N2'] = 'Tabla auxiliar — Por Estado'
    ws['N2'].font = label_font
    ws['N3'] = 'Estado'; ws['O3'] = 'Cantidad'
    for i, estado in enumerate(ESTADOS, start=4):
        ws.cell(row=i, column=14, value=estado)  # N
        ws.cell(row=i, column=15, value=f'=COUNTIF(Leads!$G$2:$G${last_row},N{i})')  # O

    # Tabla 2: por Fuente -> columnas N:O, filas 12..
    base2 = 12
    ws.cell(row=base2 - 1, column=14, value='Tabla auxiliar — Por Fuente').font = label_font
    ws.cell(row=base2, column=14, value='Fuente'); ws.cell(row=base2, column=15, value='Cantidad')
    for i, fuente in enumerate(FUENTES, start=base2 + 1):
        ws.cell(row=i, column=14, value=fuente)
        ws.cell(row=i, column=15, value=f'=COUNTIF(Leads!$H$2:$H${last_row},N{i})')

    # Tabla 3: por Producto -> columnas Q:R, filas 3..
    ws.cell(row=2, column=17, value='Tabla auxiliar — Por Producto').font = label_font
    ws.cell(row=3, column=17, value='Producto'); ws.cell(row=3, column=18, value='Cantidad')
    for i, producto in enumerate(PRODUCTOS, start=4):
        ws.cell(row=i, column=17, value=producto)
        ws.cell(row=i, column=18, value=f'=COUNTIF(Leads!$E$2:$E${last_row},Q{i})')

    # Tabla 4: Valor por Mes -> columnas Q:R, filas 16..
    base4 = 16
    ws.cell(row=base4 - 1, column=17, value='Tabla auxiliar — Valor por Mes').font = label_font
    ws.cell(row=base4, column=17, value='Mes'); ws.cell(row=base4, column=18, value='Valor USD')
    for i, mes in enumerate(MESES, start=base4 + 1):
        ws.cell(row=i, column=17, value=mes)
        # SUMPRODUCT+TEXT por si la columna Mes no calza exacto; usamos SUMIFS directo ya que Mes es texto
        ws.cell(row=i, column=18, value=f'=SUMIFS(Leads!$F$2:$F${last_row},Leads!$J$2:$J${last_row},Q{i})')
        ws.cell(row=i, column=18).number_format = '$#,##0.00'

    # ---- Gráficos nativos, bien espaciados verticalmente para no sobreponerse ----
    # Dona: embudo por estado
    chart1 = DoughnutChart()
    chart1.title = 'Embudo por Estado'
    data1 = Reference(ws, min_col=15, min_row=3, max_row=3 + len(ESTADOS))
    cats1 = Reference(ws, min_col=14, min_row=4, max_row=3 + len(ESTADOS))
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats1)
    chart1.width = 11
    chart1.height = 8
    ws.add_chart(chart1, 'A7')

    # Barras: por fuente
    chart2 = BarChart()
    chart2.type = 'col'
    chart2.title = 'Leads por Fuente'
    data2 = Reference(ws, min_col=15, min_row=base2, max_row=base2 + len(FUENTES))
    cats2 = Reference(ws, min_col=14, min_row=base2 + 1, max_row=base2 + len(FUENTES))
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    chart2.width = 11
    chart2.height = 8
    ws.add_chart(chart2, 'A25')

    # Dona: por producto
    chart3 = DoughnutChart()
    chart3.title = 'Leads por Producto'
    data3 = Reference(ws, min_col=18, min_row=3, max_row=3 + len(PRODUCTOS))
    cats3 = Reference(ws, min_col=17, min_row=4, max_row=3 + len(PRODUCTOS))
    chart3.add_data(data3, titles_from_data=True)
    chart3.set_categories(cats3)
    chart3.width = 11
    chart3.height = 8
    ws.add_chart(chart3, 'H7')

    # Barras: valor por mes
    chart4 = BarChart()
    chart4.type = 'col'
    chart4.title = 'Valor Generado por Mes'
    data4 = Reference(ws, min_col=18, min_row=base4, max_row=base4 + len(MESES))
    cats4 = Reference(ws, min_col=17, min_row=base4 + 1, max_row=base4 + len(MESES))
    chart4.add_data(data4, titles_from_data=True)
    chart4.set_categories(cats4)
    chart4.width = 11
    chart4.height = 8
    ws.add_chart(chart4, 'H25')

    for i in range(1, 20):
        ws.column_dimensions[get_column_letter(i)].width = 13

    return ws


def main():
    filas = generar_filas(CANTIDAD)
    wb = Workbook()
    ws_leads, last_row = construir_hoja_leads(wb, filas)
    construir_hoja_dashboard(wb, last_row)
    out = 'Kitty_Boutique_Sheets.xlsx'
    wb.save(out)
    print(f'Listo: {out} generado con {len(filas)} filas en "Leads" y fórmulas en "Dashboard".')
    print('AVISO: no se pudo previsualizar visualmente (no hay LibreOffice/Excel instalado en este entorno).')
    print('Se usó espaciado conservador entre tablas auxiliares (columnas N-R) y gráficos (filas 7/25), pero revisa el resultado al abrirlo.')


if __name__ == '__main__':
    main()
